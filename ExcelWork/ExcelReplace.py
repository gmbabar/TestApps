#!/usr/bin/python3
import openpyxl
# pip install openpyxl
from datetime import datetime
# pip install datetime

if __name__=="__main__":
    logFile = open("Tool.log", "ab")
    dT = datetime.now().strftime("%Y:%m:%d @ %H:%M:%S").encode()
    logFile.write(f"{dT}\n".encode())
    logFile.flush()

    fileName = input("Enter The File Name To Read : ")
    dataframe = openpyxl.load_workbook(fileName)
    dataframe1 = dataframe.active
    sheet2 = dataframe["Sheet2"]
    dataMap = {}
    for row in sheet2.iter_rows(2, sheet2.max_row):
        dataMap[str(row[0].value)] = row[1].value
    Missing = []

    for row in dataframe1.iter_rows(4, dataframe1.max_row):
        name = dataMap.get(row[3].value)
        if name:
            row[3].value = name
            print(name)
        else:
            logFile.write(f"No name found for posid : {row[3].value}\n".encode())
            logFile.flush()
            row[3].value = "Null"
    dataframe.save(fileName)
